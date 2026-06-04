import argparse
import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


SOURCE_FILE = r"F:\系统项目录入汇报\2025年入井材料统计2025.12.31-转换后.xlsx"
OUTPUT_SQL = "sql/well_project_sample_20251231_import.sql"
ISSUES_JSON = "sql/well_project_sample_20251231_issues.json"
TARGET_TABLE = "dbo.well_project_sample"
YEAR_VALUE = 2025
SKIP_SHEETS = {"年度汇总表", "统计表"}


def to_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text in {"", "#REF!", "#N/A"}:
        return None
    return text


def to_date_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(value).strip()
    if text in {"", "/", "#REF!", "#N/A"}:
        return None
    text = text.replace("/", ".")
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y.%m.%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def to_int(value):
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.strip() in {"", "#REF!", "#N/A"}:
        return None
    try:
        number = float(value)
        if number.is_integer():
            return int(number)
        return int(round(number))
    except Exception:
        return None


def sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "N'" + str(value).replace("'", "''") + "'"


def pick(row_map, *keys):
    for key in keys:
        if key in row_map and row_map[key] is not None:
            return row_map[key]
    return None


def normalize_headers(values):
    headers = []
    for value in values:
        text = to_text(value)
        headers.append(text)
    return headers


def is_meaningful_row(base):
    return any(
        base.get(field) is not None
        for field in ("sample_no", "raw_sample", "sample_name", "quantity")
    )


def is_detail_row(base):
    sample_name = base.get("sample_name")
    sample_no = base.get("sample_no")
    if sample_no in {"检测编号", "项目编号"}:
        return False
    if not sample_no:
        return False
    if not sample_name:
        return False
    return True


def load_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    invalid_rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        headers = normalize_headers(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        if not any(headers):
            continue

        prev_month = None
        prev_receive_date = None
        prev_type = None

        for row_idx, values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if all(value is None for value in values):
                continue

            row_map = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}

            month = to_text(pick(row_map, "月份"))
            receive_date = to_date_text(pick(row_map, "接样日期", "取样日期"))
            sample_type = to_text(pick(row_map, "类别", "样品类别"))

            if month is None:
                month = prev_month
            else:
                prev_month = month

            if receive_date is None:
                receive_date = prev_receive_date
            else:
                prev_receive_date = receive_date

            if sample_type is None:
                sample_type = prev_type
            else:
                prev_type = sample_type

            base = {
                "month": month,
                "receive_date": receive_date,
                "type": sample_type,
                "sample_no": to_text(pick(row_map, "检测编号", "项目编号")),
                "raw_sample": to_text(pick(row_map, "样品编号")),
                "sample_name": to_text(pick(row_map, "样品名称")),
                "quantity": to_int(pick(row_map, "数量（个）", "数量/份数", "数量")),
                "remark": to_text(pick(row_map, "备注")),
                "source": sheet_name.strip(),
                "year": YEAR_VALUE,
            }

            if not is_meaningful_row(base):
                continue

            if not is_detail_row(base):
                continue

            rows.append(base)

    return rows, invalid_rows


def build_sql(rows):
    columns = [
        "month",
        "receive_date",
        "type",
        "sample_no",
        "raw_sample",
        "sample_name",
        "quantity",
        "remark",
        "source",
        "[year]",
    ]
    row_keys = [
        "month",
        "receive_date",
        "type",
        "sample_no",
        "raw_sample",
        "sample_name",
        "quantity",
        "remark",
        "source",
        "year",
    ]
    sql = [
        "SET NOCOUNT ON;",
        "BEGIN TRANSACTION;",
        f"""
IF OBJECT_ID(N'{TARGET_TABLE}', N'U') IS NULL
BEGIN
    CREATE TABLE {TARGET_TABLE} (
        id BIGINT IDENTITY(1,1) NOT NULL PRIMARY KEY,
        [month] NVARCHAR(20) NULL,
        receive_date DATE NULL,
        [type] NVARCHAR(100) NULL,
        sample_no NVARCHAR(100) NULL,
        raw_sample NVARCHAR(255) NULL,
        sample_name NVARCHAR(255) NULL,
        quantity INT NULL,
        remark NVARCHAR(1000) NULL,
        source NVARCHAR(100) NOT NULL,
        [year] SMALLINT NOT NULL
    );
END;
""".strip(),
        f"DELETE FROM {TARGET_TABLE} WHERE [year] = {YEAR_VALUE};",
    ]
    for row in rows:
        values = ", ".join(sql_literal(row[key]) for key in row_keys)
        sql.append(f"INSERT INTO {TARGET_TABLE} ({', '.join(columns)}) VALUES ({values});")
    sql.extend(
        [
            "COMMIT TRANSACTION;",
            f"SELECT source, COUNT(1) AS total_rows FROM {TARGET_TABLE} WHERE [year] = {YEAR_VALUE} GROUP BY source ORDER BY source;",
        ]
    )
    return "\n".join(sql)


def main():
    parser = argparse.ArgumentParser(description="Generate SQL for importing well project sample Excel data.")
    parser.add_argument("--source", default=SOURCE_FILE)
    parser.add_argument("--output", default=OUTPUT_SQL)
    parser.add_argument("--issues-output", default=ISSUES_JSON)
    args = parser.parse_args()

    rows, invalid_rows = load_rows(args.source)
    issues_path = Path(args.issues_output)
    issues_path.write_text(json.dumps(invalid_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    if invalid_rows:
        print(f"source={args.source}")
        print(f"issues_output={issues_path.resolve()}")
        print(f"invalid_rows={len(invalid_rows)}")
        for item in invalid_rows[:20]:
            print(f"invalid_row={json.dumps(item, ensure_ascii=False)}")
        raise SystemExit(1)

    sql = build_sql(rows)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding="utf-8")

    print(f"source={args.source}")
    print(f"output={output_path.resolve()}")
    print(f"issues_output={issues_path.resolve()}")
    print(f"rows_to_import={len(rows)}")
    print("invalid_rows=0")


if __name__ == "__main__":
    main()
