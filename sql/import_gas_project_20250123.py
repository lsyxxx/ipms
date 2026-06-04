import argparse
import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


SOURCE_FILE = r"F:\系统项目录入汇报\ABT-2025年采气厂-水、气、硫化氢-数据汇总表-截止01.23.xlsx"
OUTPUT_SQL = "sql/gas_project_20250123_import.sql"
ISSUES_JSON = "sql/gas_project_20250123_issues.json"
TARGET_TABLE = "dbo.gas_project"
YEAR_VALUE = 2025


def to_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def to_date_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except Exception:
            return None
    text = str(value).strip()
    if text in {"", "/"}:
        return None
    return text or None


def to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "N'" + str(value).replace("'", "''") + "'"


def infer_flags(test_item):
    text = to_text(test_item) or ""
    water = 1 if "水" in text else None
    gas_component = 1 if any(key in text for key in ["气", "组分", "氦", "解吸", "解析"]) else None
    h2s = 1 if "硫化氢" in text else None
    return water, gas_component, h2s


def assign_check_module(row):
    text = row.get("test_item") or ""
    check_module_id = None
    check_module_name = None

    if row.get("water") == 1:
        if "碎屑岩" in text:
            check_module_id = "000090"
            check_module_name = "碎屑岩油藏注水水质"
        else:
            check_module_id = "000088"
            check_module_name = "水质分析"

    if row.get("h2s") == 1:
        check_module_id = "000104"
        check_module_name = "硫化氢"

    if row.get("gas_component") == 1:
        if "氦气含量测定" in text:
            check_module_id = "000201"
            check_module_name = "氦气含气量测定"
        else:
            check_module_id = "000093"
            check_module_name = "天然气组分"

    row["check_module_id"] = check_module_id
    row["check_module_name"] = check_module_name
    return row


def load_sheet_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    invalid_rows = []

    merged = wb["数据-合并"]
    merged_headers = [to_text(v) for v in next(merged.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row_idx, values in enumerate(merged.iter_rows(min_row=3, values_only=True), start=3):
        if all(value is None for value in values):
            continue
        row_map = {merged_headers[i]: values[i] for i in range(min(len(merged_headers), len(values))) if merged_headers[i]}
        base = {
            "seq_no": None,
            "receive_date": to_date_text(row_map.get("receive_date")),
            "dd_no": to_text(row_map.get("dd_no")),
            "entrust_id": to_text(row_map.get("entrust_id")),
            "entrust_id_std": to_text(row_map.get("entrust_id_std")),
            "sample_no": to_text(row_map.get("sample_no")),
            "test_item": to_text(row_map.get("test_item")),
            "well_area": to_text(row_map.get("well_area")),
            "well_type": to_text(row_map.get("well_type")),
            "well_no": to_text(row_map.get("well_no")),
            "client_name": to_text(row_map.get("client_name")),
            "finish_status": to_text(row_map.get("finish_status")),
            "remark": to_text(row_map.get("remark")),
            "water": to_int(row_map.get("water")),
            "gas_component": to_int(row_map.get("gas_component")),
            "h2s": to_int(row_map.get("h2s")),
            "is_test": to_int(row_map.get("is_test")),
            "year": YEAR_VALUE,
        }
        if base["water"] is None and base["gas_component"] is None and base["h2s"] is None:
            base["water"], base["gas_component"], base["h2s"] = infer_flags(base["test_item"])
        missing = [field for field in ("entrust_id", "sample_no", "test_item") if not base.get(field)]
        if missing:
            invalid_rows.append({"sheet": "数据-合并", "row_index": row_idx, "missing_fields": missing, **base})
            continue
        rows.append(assign_check_module(base))

    social = wb["社会样品-处理"]
    social_headers = [to_text(v) for v in next(social.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row_idx, values in enumerate(social.iter_rows(min_row=3, values_only=True), start=3):
        if all(value is None for value in values):
            continue
        row_map = {social_headers[i]: values[i] for i in range(min(len(social_headers), len(values))) if social_headers[i]}
        base = {
            "seq_no": None,
            "receive_date": to_date_text(row_map.get("receive_date")),
            "dd_no": None,
            "entrust_id": to_text(row_map.get("entrust_id")),
            "entrust_id_std": to_text(row_map.get("entrust_id_std")),
            "sample_no": to_text(row_map.get("sample_no")),
            "test_item": to_text(row_map.get("test_item")),
            "well_area": None,
            "well_type": None,
            "well_no": to_text(row_map.get("well_no")),
            "client_name": to_text(row_map.get("client_name")),
            "finish_status": to_text(row_map.get("finish_status")),
            "remark": to_text(row_map.get("remark")),
            "is_test": to_int(row_map.get("is_test")),
            "year": YEAR_VALUE,
        }
        base["water"], base["gas_component"], base["h2s"] = infer_flags(base["test_item"])
        missing = [field for field in ("entrust_id", "sample_no", "test_item") if not base.get(field)]
        if missing:
            invalid_rows.append({"sheet": "社会样品-处理", "row_index": row_idx, "missing_fields": missing, **base})
            continue
        rows.append(assign_check_module(base))

    return rows, invalid_rows


def build_sql(rows):
    row_keys = [
        "seq_no",
        "receive_date",
        "dd_no",
        "entrust_id",
        "entrust_id_std",
        "sample_no",
        "test_item",
        "well_area",
        "well_type",
        "well_no",
        "client_name",
        "finish_status",
        "remark",
        "water",
        "gas_component",
        "h2s",
        "is_test",
        "year",
        "check_module_id",
        "check_module_name",
    ]
    sql_columns = [
        "seq_no",
        "receive_date",
        "dd_no",
        "entrust_id",
        "entrust_id_std",
        "sample_no",
        "test_item",
        "well_area",
        "well_type",
        "well_no",
        "client_name",
        "finish_status",
        "remark",
        "water",
        "gas_component",
        "h2s",
        "is_test",
        "[year]",
        "check_module_id",
        "check_module_name",
    ]
    sql = [
        "SET NOCOUNT ON;",
        "BEGIN TRANSACTION;",
        f"IF COL_LENGTH(N'{TARGET_TABLE}', N'year') IS NULL ALTER TABLE {TARGET_TABLE} ADD [year] SMALLINT NULL;",
        f"IF COL_LENGTH(N'{TARGET_TABLE}', N'check_module_id') IS NULL ALTER TABLE {TARGET_TABLE} ADD check_module_id VARCHAR(20) NULL;",
        f"IF COL_LENGTH(N'{TARGET_TABLE}', N'check_module_name') IS NULL ALTER TABLE {TARGET_TABLE} ADD check_module_name VARCHAR(100) NULL;",
        f"DELETE FROM {TARGET_TABLE} WHERE [year] = {YEAR_VALUE};",
    ]
    for row in rows:
        values = ", ".join(sql_literal(row[column]) for column in row_keys)
        sql.append(f"INSERT INTO {TARGET_TABLE} ({', '.join(sql_columns)}) VALUES ({values});")
    sql.extend(
        [
            "COMMIT TRANSACTION;",
            f"SELECT [year], COUNT(1) AS total_rows FROM {TARGET_TABLE} WHERE [year] = {YEAR_VALUE} GROUP BY [year];",
        ]
    )
    return "\n".join(sql)


def main():
    parser = argparse.ArgumentParser(description="Generate SQL for importing gas_project 2025 data.")
    parser.add_argument("--source", default=SOURCE_FILE)
    parser.add_argument("--output", default=OUTPUT_SQL)
    parser.add_argument("--issues-output", default=ISSUES_JSON)
    args = parser.parse_args()

    rows, invalid_rows = load_sheet_rows(args.source)
    issues_path = Path(args.issues_output)
    issues_path.write_text(json.dumps(invalid_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    if invalid_rows:
        print(f"source={args.source}")
        print(f"issues_output={issues_path.resolve()}")
        print(f"invalid_rows={len(invalid_rows)}")
        for item in invalid_rows[:20]:
            print(f"invalid_row={json.dumps(item, ensure_ascii=False)}")
        raise SystemExit(1)

    sql = build_sql(rows)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding="utf-8")

    print(f"source={args.source}")
    print(f"output={output_path.resolve()}")
    print(f"issues_output={issues_path.resolve()}")
    print(f"rows_to_import={len(rows)}")
    print("invalid_rows=0")


if __name__ == "__main__":
    main()
