import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE_FILE = r"C:\Users\Administrator\Desktop\ABT-2026年采气厂-水、气、硫化氢-数据汇总表-截止2026.06.05.xlsx"
OUTPUT_SQL = "sql/gas_project_20260605_import.sql"
ISSUES_JSON = "sql/gas_project_20260605_issues.json"
TARGET_TABLE = "dbo.gas_project"
YEAR_VALUE = 2026


def to_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def to_date_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text in {"", "/"}:
        return None
    return text or None


def to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def std_entrust_id(value):
    text = to_text(value)
    if not text:
        return None
    return re.sub(r"[A-Za-z]$", "", text)


def derive_entrust_id_from_sample_no(sample_no):
    text = to_text(sample_no)
    if not text:
        return None
    match = re.match(r"^(.*?)[A-Za-z]\d+$", text)
    if match:
        return match.group(1)
    return None


def normalize_entrust_id(raw_entrust_id, sample_no):
    text = to_text(raw_entrust_id)
    if text and not text.startswith("="):
        return text
    return derive_entrust_id_from_sample_no(sample_no)


def infer_flags(test_item):
    text = to_text(test_item) or ""
    water = 1 if "水" in text else None
    gas_component = 1 if any(key in text for key in ["气", "组分", "氦", "解吸", "解析"]) else None
    hydrogen_sulfide = 1 if "硫化氢" in text else None
    return water, gas_component, hydrogen_sulfide


def sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "N'" + str(value).replace("'", "''") + "'"


def fill_merged_cells(ws):
    for merged in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged.bounds
        value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged))
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if ws.cell(row_idx, col_idx).value is None:
                    ws.cell(row_idx, col_idx).value = value


def is_numeric_seq(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def load_rows(workbook_path, save_unmerged):
    wb_edit = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    for sheet_name in ["采气厂样品", "社会样品"]:
        fill_merged_cells(wb_edit[sheet_name])
    if save_unmerged:
        wb_edit.save(workbook_path)
    wb_edit.close()

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    invalid_rows = []
    rows = []

    plant = wb["采气厂样品"]
    for row_idx, row in enumerate(plant.iter_rows(min_row=4, max_col=21, values_only=True), start=4):
        if not is_numeric_seq(row[0]):
            continue
        sample_no = to_text(row[6])
        base = {
            "seq_no": to_int(row[0]),
            "receive_date": to_date_text(row[1]),
            "dd_no": to_text(row[3]),
            "entrust_id": normalize_entrust_id(row[5], sample_no),
            "sample_no": sample_no,
            "test_item": to_text(row[7]),
            "well_area": to_text(row[8]),
            "well_type": to_text(row[9]),
            "well_no": to_text(row[10]),
            "client_name": to_text(row[11]),
            "finish_status": to_text(row[12]),
            "remark": to_text(row[13]),
            "water": to_int(row[14]),
            "gas_component": to_int(row[15]),
            "hydrogen_sulfide": to_int(row[16]),
            "year": YEAR_VALUE,
        }
        missing = [field for field in ("entrust_id", "sample_no", "test_item") if not base[field]]
        if missing:
            invalid_rows.append({"sheet": "采气厂样品", "row_index": row_idx, "missing_fields": missing, **base})
            continue
        if base["water"] is None and base["gas_component"] is None and base["hydrogen_sulfide"] is None:
            water, gas_component, hydrogen_sulfide = infer_flags(base["test_item"])
            base["water"] = water
            base["gas_component"] = gas_component
            base["hydrogen_sulfide"] = hydrogen_sulfide
        base["entrust_id_std"] = std_entrust_id(base["entrust_id"])
        rows.append(base)

    social = wb["社会样品"]
    for row_idx, row in enumerate(social.iter_rows(min_row=3, max_col=9, values_only=True), start=3):
        if not is_numeric_seq(row[0]):
            continue
        sample_no = to_text(row[3])
        water, gas_component, hydrogen_sulfide = infer_flags(row[4])
        base = {
            "seq_no": to_int(row[0]),
            "receive_date": to_date_text(row[1]),
            "dd_no": None,
            "entrust_id": normalize_entrust_id(row[2], sample_no),
            "sample_no": sample_no,
            "test_item": to_text(row[4]),
            "well_area": None,
            "well_type": None,
            "well_no": to_text(row[5]),
            "client_name": to_text(row[6]),
            "finish_status": to_text(row[7]),
            "remark": to_text(row[8]),
            "water": water,
            "gas_component": gas_component,
            "hydrogen_sulfide": hydrogen_sulfide,
            "year": YEAR_VALUE,
            "entrust_id_std": std_entrust_id(normalize_entrust_id(row[2], sample_no)),
        }
        missing = [field for field in ("entrust_id", "sample_no", "test_item") if not base[field]]
        if missing:
            invalid_rows.append({"sheet": "社会样品", "row_index": row_idx, "missing_fields": missing, **base})
            continue
        rows.append(base)

    return rows, invalid_rows


def build_sql(rows):
    row_keys = [
        "seq_no",
        "receive_date",
        "dd_no",
        "entrust_id",
        "entrust_id_std",
        "sample_no",
        "test_item",
        "well_area",
        "well_type",
        "well_no",
        "client_name",
        "finish_status",
        "remark",
        "water",
        "gas_component",
        "hydrogen_sulfide",
        "year",
    ]
    sql_columns = [
        "seq_no",
        "receive_date",
        "dd_no",
        "entrust_id",
        "entrust_id_std",
        "sample_no",
        "test_item",
        "well_area",
        "well_type",
        "well_no",
        "client_name",
        "finish_status",
        "remark",
        "water",
        "gas_component",
        "hydrogen_sulfide",
        "[year]",
    ]
    sql = [
        "SET NOCOUNT ON;",
        "BEGIN TRANSACTION;",
        f"IF COL_LENGTH(N'{TARGET_TABLE}', N'year') IS NULL ALTER TABLE {TARGET_TABLE} ADD [year] INT NULL;",
        f"UPDATE {TARGET_TABLE} SET [year] = CASE WHEN entrust_id LIKE N'%2025%' THEN 2025 WHEN entrust_id LIKE N'%2026%' THEN 2026 ELSE [year] END;",
        f"DELETE FROM {TARGET_TABLE} WHERE [year] = {YEAR_VALUE};",
    ]
    for row in rows:
        values = ", ".join(sql_literal(row[column]) for column in row_keys)
        sql.append(
            f"INSERT INTO {TARGET_TABLE} ({', '.join(sql_columns)}) VALUES ({values});"
        )
    sql.extend(
        [
            "COMMIT TRANSACTION;",
            f"SELECT [year], COUNT(1) AS total_rows FROM {TARGET_TABLE} GROUP BY [year] ORDER BY [year];",
        ]
    )
    return "\n".join(sql)


def main():
    parser = argparse.ArgumentParser(description="Unmerge Excel and generate SQL for gas_project 2026 import.")
    parser.add_argument("--source", default=SOURCE_FILE)
    parser.add_argument("--output", default=OUTPUT_SQL)
    parser.add_argument("--issues-output", default=ISSUES_JSON)
    parser.add_argument("--no-save-unmerged", action="store_true", help="Do not save the unmerged workbook back to source file")
    args = parser.parse_args()

    rows, invalid_rows = load_rows(args.source, save_unmerged=not args.no_save_unmerged)
    issues_path = Path(args.issues_output)
    issues_path.write_text(json.dumps(invalid_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    if invalid_rows:
        print(f"source={args.source}")
        print(f"issues_output={issues_path.resolve()}")
        print(f"invalid_rows={len(invalid_rows)}")
        for item in invalid_rows[:20]:
            print(f"invalid_row={json.dumps(item, ensure_ascii=False)}")
        raise SystemExit(1)

    sql = build_sql(rows)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding="utf-8")

    print(f"source={args.source}")
    print(f"output={output_path.resolve()}")
    print(f"issues_output={issues_path.resolve()}")
    print(f"rows_to_import={len(rows)}")
    print("invalid_rows=0")


if __name__ == "__main__":
    main()
