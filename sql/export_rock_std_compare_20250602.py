import json
import re
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(r"E:\github\ipms")
STD_JSON = BASE_DIR / "sql" / "rock_project_std_20250602.json"
CHECKMODULE_JSON = BASE_DIR / "sql" / "t_checkmodule_20250602.json"
OUTPUT_XLSX = BASE_DIR / "sql" / "rock_project_std_compare_20250602.xlsx"


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


def normalize(text):
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = re.sub(r"[\s\u3000]+", "", text)
    text = re.sub(r"[（）()【】\[\]，,、；;：:·/\\]+", "", text)
    return text


def exact_fname_match(std, modules):
    for module in modules:
        if std == module["Fname"]:
            return module, "sql_exact_fname"
    return None, None


def best_unmatched_supplement(std, modules):
    norm_std = normalize(std)
    for module in modules:
        if norm_std and module.get("BreName") and norm_std == normalize(module["BreName"]):
            return module, "breName_normalized"
    return None, None


def score_candidate(std, module):
    fname = module["Fname"] or ""
    brename = module.get("BreName") or ""
    norm_std = normalize(std)
    norm_fname = normalize(fname)
    norm_brename = normalize(brename)
    candidates = [(fname, norm_fname, "fname"), (brename, norm_brename, "breName")]

    best_score = -1
    best_basis = ""
    best_name = ""
    for raw_name, norm_name, source in candidates:
        if not raw_name or not norm_name:
            continue
        score = 0
        basis = source
        if norm_std and norm_name and norm_std in norm_name:
            score = 95
            basis = f"{source}_contains_std"
        elif norm_std and norm_name and norm_name in norm_std and len(norm_name) >= 4:
            score = 85
            basis = f"{source}_contained_by_std"
        else:
            ratio = SequenceMatcher(None, norm_std, norm_name).ratio()
            score = int(round(ratio * 100))
            basis = f"{source}_similarity_{score}"
        if score > best_score or (score == best_score and len(raw_name) < len(best_name or "ZZZ")):
            best_score = score
            best_basis = basis
            best_name = raw_name
    return best_score, best_basis, best_name


def build_rows(std_rows, modules):
    rows = []
    for item in std_rows:
        std = item["std"]
        module, match_basis = exact_fname_match(std, modules)
        if module:
            rows.append(
                {
                    "std": std,
                    "fid": module["Fid"],
                    "fname": module["Fname"],
                    "match_basis": match_basis,
                    "possible_standard_name": None,
                    "possible_standard_fid": None,
                    "possible_match_basis": None,
                }
            )
            continue

        supplement, supplement_basis = best_unmatched_supplement(std, modules)
        if supplement:
            rows.append(
                {
                    "std": std,
                    "fid": None,
                    "fname": None,
                    "match_basis": "unmatched_by_sql",
                    "possible_standard_name": supplement["Fname"],
                    "possible_standard_fid": supplement["Fid"],
                    "possible_match_basis": supplement_basis,
                }
            )
            continue

        scored = []
        for module in modules:
            score, basis, best_name = score_candidate(std, module)
            if score >= 45:
                scored.append((score, basis, best_name, module))
        scored.sort(key=lambda x: (-x[0], len(x[2]), x[3]["Fid"]))
        top = scored[:3]

        rows.append(
            {
                "std": std,
                "fid": None,
                "fname": None,
                "match_basis": "unmatched_by_sql",
                "possible_standard_name": "；".join(t[2] for t in top) if top else None,
                "possible_standard_fid": "；".join(t[3]["Fid"] for t in top) if top else None,
                "possible_match_basis": "；".join(f"{t[1]}({t[0]})" for t in top) if top else None,
            }
        )
    return rows


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 60)


def main():
    std_rows = load_json(STD_JSON)
    modules = load_json(CHECKMODULE_JSON)
    rows = build_rows(std_rows, modules)

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    headers = [
        "std",
        "fid",
        "fname",
        "match_basis",
        "possible_standard_name",
        "possible_standard_fid",
        "possible_match_basis",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    auto_width(ws)

    summary = wb.create_sheet("summary")
    summary.append(["match_basis", "count"])
    stats = {}
    for row in rows:
        stats[row["match_basis"]] = stats.get(row["match_basis"], 0) + 1
    for key, value in sorted(stats.items()):
        summary.append([key, value])
    summary.append([])
    summary.append(["total", len(rows)])
    summary.append(["sql_exact_fname_match", stats.get("sql_exact_fname", 0)])
    summary.append(["unmatched_by_sql", stats.get("unmatched_by_sql", 0)])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    auto_width(summary)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"output={OUTPUT_XLSX}")
    print(f"rows={len(rows)}")
    print(f"unmatched={stats.get('unmatched', 0)}")


if __name__ == "__main__":
    main()
