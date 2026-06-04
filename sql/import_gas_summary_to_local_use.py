import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


TABLE_NAME = "dbo.gas_summary_20250123"
SOURCE_FILE = r"C:\Users\Administrator\Desktop\ABT-2025年采气厂-水、气、硫化氢-数据汇总表-截止01.23.xlsx"


def to_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def to_date_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except Exception:
            return None
    text = str(value).strip()
    if text in {"", "/"}:
        return None
    return text or None


def to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def std_entrust_id(value):
    text = to_text(value)
    if not text:
        return None
    return re.sub(r"[A-Za-z]$", "", text)


def infer_flags(test_item):
    text = to_text(test_item) or ""
    water = 1 if "水" in text else None
    gas_component = 1 if any(key in text for key in ["气", "组分", "氦", "解吸", "解析"]) else None
    hydrogen_sulfide = 1 if "硫化氢" in text else None
    return water, gas_component, hydrogen_sulfide


def sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "N'" + str(value).replace("'", "''") + "'"


def missing_required_fields(base):
    missing = []
    for field in ("entrust_id", "sample_no", "test_item"):
        if not base.get(field):
            missing.append(field)
    return missing


def load_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    invalid_rows = []

    ws = wb["数据-合并"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_col=15, values_only=True), start=3):
        base = {
            "seq_no": to_int(row[0]),
            "receive_date": to_date_text(row[1]),
            "dd_no": to_text(row[2]),
            "entrust_id": to_text(row[3]),
            "sample_no": to_text(row[4]),
            "test_item": to_text(row[5]),
            "well_area": to_text(row[6]),
            "well_type": to_text(row[7]),
            "well_no": to_text(row[8]),
            "client_name": to_text(row[9]),
            "finish_status": to_text(row[10]),
            "remark": to_text(row[11]),
            "water": to_int(row[12]),
            "gas_component": to_int(row[13]),
            "hydrogen_sulfide": to_int(row[14]),
        }
        missing = missing_required_fields(base)
        if missing:
            invalid_rows.append({
                "sheet": "数据-合并",
                "row_index": row_idx,
                "missing_fields": missing,
                **base,
            })
            continue
        if base["water"] is None and base["gas_component"] is None and base["hydrogen_sulfide"] is None:
            inferred = infer_flags(base["test_item"])
            base["water"] = inferred[0]
            base["gas_component"] = inferred[1]
            base["hydrogen_sulfide"] = inferred[2]
        base["std_entrust_id"] = std_entrust_id(base["entrust_id"])
        rows.append(base)

    ws = wb["社会样品-处理"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_col=10, values_only=True), start=3):
        water, gas_component, hydrogen_sulfide = infer_flags(row[5])
        base = {
            "seq_no": to_int(row[0]),
            "receive_date": to_date_text(row[2]),
            "dd_no": None,
            "entrust_id": to_text(row[3]),
            "sample_no": to_text(row[4]),
            "test_item": to_text(row[5]),
            "well_area": None,
            "well_type": None,
            "well_no": to_text(row[6]),
            "client_name": to_text(row[7]),
            "finish_status": to_text(row[8]),
            "remark": to_text(row[9]),
            "water": water,
            "gas_component": gas_component,
            "hydrogen_sulfide": hydrogen_sulfide,
            "std_entrust_id": std_entrust_id(row[3]),
        }
        missing = missing_required_fields(base)
        if missing:
            invalid_rows.append({
                "sheet": "社会样品-处理",
                "row_index": row_idx,
                "missing_fields": missing,
                **base,
            })
            continue
        rows.append(base)

    return rows, invalid_rows


def build_sql(rows, table_name):
    columns = [
        "seq_no",
        "receive_date",
        "dd_no",
        "entrust_id",
        "std_entrust_id",
        "sample_no",
        "test_item",
        "well_area",
        "well_type",
        "well_no",
        "client_name",
        "finish_status",
        "remark",
        "water",
        "gas_component",
        "hydrogen_sulfide",
    ]
    sql = [
        "SET NOCOUNT ON;",
        "BEGIN TRANSACTION;",
        f"IF OBJECT_ID(N'{table_name}', N'U') IS NOT NULL DROP TABLE {table_name};",
        f"""
CREATE TABLE {table_name} (
    id BIGINT IDENTITY(1,1) NOT NULL PRIMARY KEY,
    seq_no INT NULL,
    receive_date DATE NULL,
    dd_no NVARCHAR(100) NULL,
    entrust_id NVARCHAR(100) NOT NULL,
    std_entrust_id NVARCHAR(100) NULL,
    sample_no NVARCHAR(100) NOT NULL,
    test_item NVARCHAR(1000) NOT NULL,
    well_area NVARCHAR(255) NULL,
    well_type NVARCHAR(255) NULL,
    well_no NVARCHAR(255) NULL,
    client_name NVARCHAR(255) NULL,
    finish_status NVARCHAR(100) NULL,
    remark NVARCHAR(1000) NULL,
    water INT NULL,
    gas_component INT NULL,
    hydrogen_sulfide INT NULL
);
""".strip(),
    ]
    for row in rows:
        values = ", ".join(sql_literal(row[column]) for column in columns)
        sql.append(
            f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({values});"
        )
    sql.extend(
        [
            "COMMIT TRANSACTION;",
            f"SELECT COUNT(1) AS imported_count FROM {table_name};",
        ]
    )
    return "\n".join(sql)


def main():
    parser = argparse.ArgumentParser(description="Generate SQL for importing gas summary Excel data.")
    parser.add_argument("--source", default=SOURCE_FILE, help="Excel source file path")
    parser.add_argument("--table", default=TABLE_NAME, help="Target table name, e.g. dbo.gas_summary_20250123")
    parser.add_argument("--output", default="sql/gas_summary_20250123_import.sql", help="Output SQL file path")
    parser.add_argument("--issues-output", default="sql/gas_summary_20250123_issues.json", help="Output JSON file for invalid rows")
    args = parser.parse_args()

    rows, invalid_rows = load_rows(args.source)
    issues_path = Path(args.issues_output)
    issues_path.write_text(json.dumps(invalid_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    if invalid_rows:
        print(f"source={args.source}")
        print(f"table={args.table}")
        print(f"issues_output={issues_path.resolve()}")
        print(f"invalid_rows={len(invalid_rows)}")
        for item in invalid_rows[:20]:
            print(f"invalid_row={json.dumps(item, ensure_ascii=False)}")
        raise SystemExit(1)

    sql = build_sql(rows, args.table)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding="utf-8")

    print(f"source={args.source}")
    print(f"table={args.table}")
    print(f"output={output_path.resolve()}")
    print(f"issues_output={issues_path.resolve()}")
    print(f"rows_to_import={len(rows)}")
    print("invalid_rows=0")


if __name__ == "__main__":
    main()
