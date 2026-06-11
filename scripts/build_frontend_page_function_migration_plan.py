from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(r"E:\github\ipms")
DOCS = ROOT / "docs"
OLD_VIEWS_ROOT = Path(r"E:\github\client_prod\src\views")
NEW_VIEWS_ROOT = Path(r"E:\github\lims\lab-lims-frontend\src\views")
NEW_API_ROOT = Path(r"E:\github\lims\lab-lims-frontend\src\api")
ABT_ROOT = Path(r"E:\github\abt-lims")

DETAIL_WORKBOOK = DOCS / "client-prod-views-page-migration-detailed.xlsx"
URL_CSV = DOCS / "client_prod-url-inventory.csv"
OUTPUT_XLSX = DOCS / "client-prod-frontend-migration-plan-by-page-function.xlsx"

EXCLUDED_CATEGORIES = {
    "低代码表单与流程设计",
    "系统管理",
    "系统配置/编码规则",
    "OidcCallback.vue",
    "OidcRedirect.vue",
    "技术页面",
    "测试页面",
}

EXCLUDED_FOLDERS = {
    "buildertables",
    "flowinstances",
    "flowschemes",
    "forms",
    "modulemanager",
    "resources",
    "dataprivilegerules",
    "openjobs",
    "orgmanager",
    "rolemanager",
    "usermanager",
    "login",
    "errorPage",
    "iframePage",
    "layout",
    "syssetting",
    "syslogs",
    "categories",
}

PAGE_PRIORITY = {
    "检验检测执行链": "P1",
    "审批中心": "P1",
    "结算管理": "P1",
    "库存管理": "P2",
    "检测基础配置": "P2",
    "统计分析/工作台": "P2",
    "小程序运营": "P2",
    "客户管理": "P2",
    "供应商管理": "P2",
    "供应商/外协管理": "P2",
    "财务管理": "P3",
    "工资/薪资": "P3",
    "车辆管理": "P3",
    "野外考勤": "P3",
    "安全检查": "P3",
    "OA/记录": "P3",
    "协议/合同/报价": "P3",
    "WMS/入库单": "P3",
    "文件与附件": "P3",
    "签名管理": "P3",
    "人事管理": "P3",
}

ACTION_ORDER = {
    "分页查询": 1,
    "列表查询": 2,
    "选项查询": 3,
    "新增": 4,
    "编辑": 5,
    "删除": 6,
    "导入": 7,
    "导出": 8,
    "详情查询": 9,
    "详情跳转": 10,
    "新增跳转": 11,
    "编辑跳转": 12,
    "验收跳转": 13,
    "审批提交": 14,
    "审批处理": 15,
    "撤销": 16,
    "重新提交": 17,
    "暂存": 18,
    "预览": 19,
    "流程记录": 20,
    "附件上传": 21,
    "复制": 22,
    "打印/PDF": 23,
    "其他": 99,
}


def normalize_rel(path: str) -> str:
    return path.replace("/", "\\").strip()


def extract_script(text: str) -> str:
    m = re.search(r"<script[^>]*>([\s\S]*?)</script>", text, re.IGNORECASE)
    return m.group(1) if m else text


def read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def route_path_to_list(value: str | None) -> List[str]:
    if not value:
        return []
    return [item.strip() for item in re.split(r"[；;]", str(value)) if item and str(item).strip() and str(item).strip() != "None"]


def build_old_api_map() -> Dict[Tuple[str, str], dict]:
    result: Dict[Tuple[str, str], dict] = {}
    with URL_CSV.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            file_path = row["frontend_file"]
            if not file_path or row.get("frontend_kind") != "api":
                continue
            key = (normalize_rel(Path(file_path).relative_to(Path(r"E:\github\client_prod\src\api")).as_posix()), row["frontend_function"])
            result[key] = row
    return result


def parse_named_imports(script: str) -> tuple[dict, dict]:
    named: dict[str, tuple[str, str]] = {}
    namespaces: dict[str, str] = {}
    for m in re.finditer(r"import\s+\{([^}]+)\}\s+from\s+['\"](@/api/[^'\"]+)['\"]", script):
        items, source = m.groups()
        rel = normalize_rel(source.replace("@/api/", "").replace(".js", "").replace(".ts", "") + Path(source).suffix)
        if rel.endswith(".js") or rel.endswith(".ts"):
            pass
        else:
            if (OLD_VIEWS_ROOT / rel).exists():
                pass
            rel = rel + ".js"
        for raw in items.split(","):
            item = raw.strip()
            if not item:
                continue
            if " as " in item:
                original, local = [x.strip() for x in item.split(" as ", 1)]
            else:
                original = local = item
            named[local] = (rel, original)
    for m in re.finditer(r"import\s+\*\s+as\s+(\w+)\s+from\s+['\"](@/api/[^'\"]+)['\"]", script):
        local, source = m.groups()
        rel = normalize_rel(source.replace("@/api/", "").replace(".js", "").replace(".ts", "") + Path(source).suffix)
        if not rel.endswith((".js", ".ts")):
            rel = rel + ".js"
        namespaces[local] = rel
    return named, namespaces


def fix_module_extension(base: str, root: Path) -> str | None:
    base = normalize_rel(base)
    candidates = [base, f"{base}.js", f"{base}.ts"]
    for candidate in candidates:
        if (root / candidate).exists():
            return normalize_rel(candidate)
    return None


def parse_named_imports_v2(script: str, root: Path) -> tuple[dict, dict]:
    named: dict[str, tuple[str, str]] = {}
    namespaces: dict[str, str] = {}
    for m in re.finditer(r"import\s+\{([^}]+)\}\s+from\s+['\"](@/api/[^'\"]+)['\"]", script):
        items, source = m.groups()
        rel = source.replace("@/api/", "")
        fixed = fix_module_extension(rel, root)
        if not fixed:
            continue
        for raw in items.split(","):
            item = raw.strip()
            if not item:
                continue
            if " as " in item:
                original, local = [x.strip() for x in item.split(" as ", 1)]
            else:
                original = local = item
            named[local] = (fixed, original)
    for m in re.finditer(r"import\s+\*\s+as\s+(\w+)\s+from\s+['\"](@/api/[^'\"]+)['\"]", script):
        local, source = m.groups()
        rel = source.replace("@/api/", "")
        fixed = fix_module_extension(rel, root)
        if fixed:
            namespaces[local] = fixed
    return named, namespaces


def scan_called_imports(script: str, named: dict, namespaces: dict) -> List[dict]:
    results: List[dict] = []
    for local, (module, original) in named.items():
        if re.search(rf"\b{re.escape(local)}\s*\(", script):
            results.append({"kind": "api", "module": module, "function": original, "call": local})
    for ns, module in namespaces.items():
        for fn in sorted(set(re.findall(rf"\b{re.escape(ns)}\.(\w+)\s*\(", script))):
            results.append({"kind": "api", "module": module, "function": fn, "call": f"{ns}.{fn}"})
    return results


def scan_helper_urls(script: str) -> List[dict]:
    rows: List[dict] = []
    for helper, url in re.findall(r"(\w+)\(\s*['\"](/[^'\"]+)['\"]", script):
        rows.append({"kind": "helper", "helper": helper, "url": url})
    return rows


def scan_router_targets(script: str) -> List[dict]:
    targets: List[dict] = []
    for target in re.findall(r"(?:\$router|router)\.push\(\s*['\"]([^'\"]+)['\"]", script):
        targets.append({"kind": "route", "target": target})
    for target in re.findall(r"path:\s*['\"]([^'\"]+)['\"]", script):
        if target.startswith("/"):
            targets.append({"kind": "route", "target": target})
    return targets


def parse_api_module_urls(path: Path) -> Dict[str, str]:
    text = read_text(path)
    mapping: Dict[str, str] = {}
    prop_urls: Dict[str, str] = {}

    for m in re.finditer(r"export function\s+(\w+)\s*\([^)]*\)\s*\{([\s\S]*?)\n\}", text):
        name, body = m.groups()
        um = re.search(r"(?:url\s*:\s*['\"]([^'\"]+)['\"]|request\.\w+\([^'\"]*['\"]([^'\"]+)['\"])", body)
        if um:
            mapping[name] = um.group(1) or um.group(2)

    for m in re.finditer(r"export const\s+(\w+)\s*=\s*\([^)]*\)\s*=>[\s\S]*?request\.\w+\([^'\"]*['\"]([^'\"]+)['\"]", text):
        mapping[m.group(1)] = m.group(2)

    for m in re.finditer(r"(\w+)\s*:\s*\([^)]*\)\s*=>[\s\S]*?request\.\w+\([^'\"]*['\"]([^'\"]+)['\"]", text):
        prop_urls[m.group(1)] = m.group(2)
    for m in re.finditer(r"(\w+)\([^)]*\)\s*\{[\s\S]*?request\.\w+\([^'\"]*['\"]([^'\"]+)['\"]", text):
        prop_urls[m.group(1)] = m.group(2)

    object_aliases: Dict[str, Dict[str, str]] = {}
    for m in re.finditer(r"export const\s+(\w+)\s*=\s*\{([\s\S]*?)\n\}", text):
        object_name, body = m.groups()
        object_aliases[object_name] = {}
        for pm in re.finditer(r"(\w+)\s*:\s*\([^)]*\)\s*=>[\s\S]*?request\.\w+\([^'\"]*['\"]([^'\"]+)['\"]", body):
            object_aliases[object_name][pm.group(1)] = pm.group(2)
        for pm in re.finditer(r"(\w+)\([^)]*\)\s*\{[\s\S]*?request\.\w+\([^'\"]*['\"]([^'\"]+)['\"]", body):
            object_aliases[object_name][pm.group(1)] = pm.group(2)

    for m in re.finditer(r"export const\s+(\w+)\s*=\s*(\w+)\.(\w+)", text):
        alias, obj, prop = m.groups()
        if obj in object_aliases and prop in object_aliases[obj]:
            mapping[alias] = object_aliases[obj][prop]
        elif prop in prop_urls:
            mapping[alias] = prop_urls[prop]

    return mapping


def build_new_api_map() -> Dict[Tuple[str, str], str]:
    result: Dict[Tuple[str, str], str] = {}
    for file in NEW_API_ROOT.rglob("*.ts"):
        rel = normalize_rel(str(file.relative_to(NEW_API_ROOT)))
        urls = parse_api_module_urls(file)
        for fn, url in urls.items():
            result[(rel, fn)] = url
    return result


def build_abt_routes() -> Dict[str, str]:
    routes: Dict[str, str] = {}
    controller_root = ABT_ROOT / "admin-api" / "src" / "main" / "java"
    for file in controller_root.rglob("*Controller.java"):
        text = read_text(file)
        base = ""
        m = re.search(r'@RequestMapping\("([^"]+)"\)', text)
        if m:
            base = m.group(1)
        for mm in re.finditer(r'@(GetMapping|PostMapping|PutMapping|DeleteMapping|PatchMapping)\("([^"]+)"\)', text):
            path = mm.group(2)
            routes[normalize_rel(base + path)] = str(file)
        for mm in re.finditer(r"@(GetMapping|PostMapping|PutMapping|DeleteMapping|PatchMapping)\s*$", text, re.MULTILINE):
            line_start = mm.start()
            snippet = text[line_start: line_start + 240]
            method_m = re.search(r"public\s+\w+[<\w, ?]*\s+(\w+)\(", snippet)
            if method_m:
                routes[normalize_rel(base)] = str(file)
    return routes


def classify_action(url: str = "", fn: str = "", route_target: str = "") -> str:
    text = f"{url} {fn} {route_target}".lower()
    if route_target:
        if any(k in route_target.lower() for k in ["/detail", "/dtl/", "detail/"]):
            return "详情跳转"
        if any(k in route_target.lower() for k in ["/apply", "/add", "/new"]):
            return "新增跳转"
        if "/edit" in route_target.lower():
            return "编辑跳转"
        if "/accept/" in route_target.lower():
            return "验收跳转"
        return "其他"
    if any(k in text for k in ["export", "download", "create/pdf", "pdf"]):
        return "导出"
    if "import" in text:
        return "导入"
    if "upload" in text:
        return "附件上传"
    if any(k in text for k in ["delete", "/del", "remove"]):
        return "删除"
    if "restart" in text:
        return "重新提交"
    if "revoke" in text:
        return "撤销"
    if "approve" in text or "delegate" in text or "resolve" in text:
        return "审批处理"
    if "accept" in text:
        return "审批处理"
    if "record" in text:
        return "流程记录"
    if "preview" in text:
        return "预览"
    if "temp" in text:
        return "暂存"
    if any(k in text for k in ["/all", "/page", "/load", "/list", "findall", "getlist", "myapply", "/todo", "/done"]):
        if any(k in text for k in ["/todo", "/done", "/myapply", "/all", "/page", "/load", "/list"]):
            return "分页查询"
        return "列表查询"
    if any(k in text for k in ["/dept", "/enabled", "/tree", "/companies", "/basic", "/query", "/find/all", "/find/active"]):
        return "选项查询"
    if any(k in text for k in ["apply", "save", "add", "create"]):
        return "新增"
    if any(k in text for k in ["update", "edit", "status", "copy"]):
        if "copy" in text:
            return "复制"
        return "编辑"
    if any(k in text for k in ["/detail", "getbyid", "/{id}", "/load"]):
        return "详情查询"
    return "其他"


def describe_feature(action: str, page_title: str, page_rel: str, url: str = "", route_target: str = "") -> str:
    subject = page_title or Path(page_rel).stem
    mapping = {
        "分页查询": f"{subject}分页查询",
        "列表查询": f"{subject}列表查询",
        "选项查询": f"{subject}下拉/辅助数据查询",
        "新增": f"{subject}新增/提交",
        "编辑": f"{subject}编辑/更新",
        "删除": f"{subject}删除",
        "导入": f"{subject}导入",
        "导出": f"{subject}导出/下载",
        "详情查询": f"{subject}详情/加载",
        "详情跳转": f"{subject}详情跳转",
        "新增跳转": f"{subject}新增页跳转",
        "编辑跳转": f"{subject}编辑页跳转",
        "验收跳转": f"{subject}验收页跳转",
        "审批提交": f"{subject}审批提交",
        "审批处理": f"{subject}审批处理",
        "撤销": f"{subject}撤销",
        "重新提交": f"{subject}重新提交",
        "暂存": f"{subject}暂存",
        "预览": f"{subject}预览",
        "流程记录": f"{subject}流程记录",
        "附件上传": f"{subject}附件上传",
        "复制": f"{subject}复制",
        "打印/PDF": f"{subject}打印/PDF",
        "其他": f"{subject}其他功能",
    }
    desc = mapping.get(action, f"{subject}{action}")
    if url:
        return f"{desc}（{url}）"
    if route_target:
        return f"{desc}（{route_target}）"
    return desc


def normalize_abt_path(url: str) -> str:
    return url if url.startswith("/api/") else f"/api{url}"


def rewrite_workflow_url(url: str) -> str | None:
    if not url.startswith("/"):
        return None

    mapping = {
        "rbs": "reimburse",
        "trip": "trip",
        "pay": "pay-voucher",
        "loan": "loan",
        "inv": "invoice",
        "invoffset": "invoice-offset",
        "purchase": "purchase",
        "wf/sbct": "subcontract-testing",
        "wf/sbctstl": "subcontract-testing-settlement",
    }

    def biz_key(u: str) -> Tuple[str | None, str]:
        if u.startswith("/wf/sbctstl/"):
            return "subcontract-testing-settlement", u[len("/wf/sbctstl/"):]
        if u.startswith("/wf/sbct/"):
            return "subcontract-testing", u[len("/wf/sbct/"):]
        parts = u.strip("/").split("/")
        if not parts:
            return None, ""
        first = parts[0]
        biz = mapping.get(first)
        if biz:
            return biz, "/".join(parts[1:])
        return None, ""

    biz, tail = biz_key(url)
    if not biz:
        return None

    if tail in {"todo", "done", "myapply", "all"}:
        return f"/api/admin/approval/{biz}/{tail}/page"
    if tail == "todo/count":
        return f"/api/admin/approval/{biz}/todo/count"
    if tail.startswith("load"):
        return f"/api/admin/approval/{biz}/{{id}}"
    if tail.startswith("record"):
        return f"/api/admin/approval/{biz}/{{id}}/record"
    if tail == "apply":
        return f"/api/admin/approval/{biz}/apply"
    if tail == "approve":
        return f"/api/admin/approval/{biz}/approve"
    if tail == "preview":
        return f"/api/admin/approval/{biz}/preview"
    if tail == "revoke":
        return f"/api/admin/approval/{biz}/{{id}}/revoke"
    if tail == "restart":
        return f"/api/admin/approval/{biz}/{{id}}/restart"
    if tail.startswith("del"):
        return f"/api/admin/approval/{biz}/{{id}}"
    if tail == "find/companies":
        return None
    if tail == "find/apply/dtl":
        return None
    if tail == "find/apply/samples":
        return None
    if tail == "find/apply/duplicated":
        return None
    if tail == "validate/duplicate":
        return None
    if tail == "samples/page":
        return None
    return None


def rewrite_purchase_special(url: str) -> str | None:
    special = {
        "/purchase/export/list": "/api/admin/approval/purchase/export-list",
        "/purchase/export/excel": "/api/admin/approval/purchase/{id}/export-excel",
        "/purchase/export/accept/pdf": "/api/admin/approval/purchase/{id}/export-accept-excel",
        "/purchase/create/accept/excel": "/api/admin/approval/purchase/{id}/export-accept-excel",
        "/purchase/accept/all": "/api/admin/approval/purchase/approve",
        "/purchase/dept": None,
        "/purchase/create/pdf": None,
        "/purchase/export/pdf": "/api/admin/approval/purchase/{id}/export-excel",
        "/mtr/find/all": None,
        "/mtr/find/active": None,
    }
    if url in special:
        return special[url]
    return rewrite_workflow_url(url)


def resolve_abt_url(url: str, abt_routes: Dict[str, str]) -> tuple[str, str]:
    if not url:
        return "", ""
    candidate = None
    if url.startswith("/admin/") or url.startswith("/auth/"):
        candidate = normalize_abt_path(url)
    elif url.startswith("/notify/"):
        candidate = normalize_abt_path(url)
    elif url.startswith("/files/"):
        candidate = normalize_abt_path(url)
    elif url.startswith("/purchase/") or url.startswith("/wf/") or any(url.startswith(f"/{x}/") for x in ["rbs", "trip", "pay", "loan", "inv", "invoffset"]):
        candidate = rewrite_purchase_special(url)
    elif url.startswith("/sys/"):
        candidate = None
    elif url.startswith("/api/"):
        candidate = url
    else:
        candidate = normalize_abt_path(url)

    if not candidate:
        return "", ""

    controller = abt_routes.get(normalize_rel(candidate), "")
    return candidate, controller


def build_new_page_candidates(new_api_map: Dict[Tuple[str, str], str], abt_routes: Dict[str, str]) -> Dict[str, List[dict]]:
    cache: Dict[str, List[dict]] = {}
    for file in NEW_VIEWS_ROOT.rglob("*.vue"):
        rel = normalize_rel(str(file.relative_to(NEW_VIEWS_ROOT)))
        script = extract_script(read_text(file))
        named, namespaces = parse_named_imports_v2(script, NEW_API_ROOT)
        entries = scan_called_imports(script, named, namespaces)
        rows: List[dict] = []
        for entry in entries:
            key = (entry["module"], entry["function"])
            url = new_api_map.get(key, "")
            action = classify_action(url=url, fn=entry["function"])
            abt_url, controller = resolve_abt_url(url, abt_routes)
            rows.append({
                "action": action,
                "new_call": f"{entry['module']}::{entry['function']}",
                "new_api_url": url,
                "abt_url": abt_url,
                "controller": controller,
            })
        for helper in scan_helper_urls(script):
            action = classify_action(url=helper["url"], fn=helper["helper"])
            abt_url, controller = resolve_abt_url(helper["url"], abt_routes)
            rows.append({
                "action": action,
                "new_call": helper["helper"],
                "new_api_url": helper["url"],
                "abt_url": abt_url,
                "controller": controller,
            })
        cache[rel] = rows
    return cache


def choose_new_match(action: str, candidates: List[dict], source_url: str = "") -> dict | None:
    same_action = [row for row in candidates if row["action"] == action]
    if source_url and same_action:
        source_tokens = [token for token in re.split(r"[/{}_-]+", source_url.lower()) if token]
        best = None
        best_score = -1
        for row in same_action:
            target_tokens = [token for token in re.split(r"[/{}_-]+", (row.get("new_api_url") or "").lower()) if token]
            score = len(set(source_tokens) & set(target_tokens))
            if score > best_score:
                best = row
                best_score = score
        if best is not None:
            return best
    if same_action:
        return same_action[0]
    fallback_map = {
        "详情跳转": "详情查询",
        "新增跳转": "新增",
        "编辑跳转": "编辑",
        "验收跳转": "审批处理",
    }
    mapped = fallback_map.get(action)
    if mapped:
        for row in candidates:
            if row["action"] == mapped:
                return row
    return candidates[0] if candidates else None


def load_page_rows() -> List[dict]:
    wb = load_workbook(DETAIL_WORKBOOK, read_only=True, data_only=True)
    ws = wb["views页面迁移清单"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    result = []
    for row in rows:
        category = row[idx["功能分类"]]
        old_file = row[idx["旧页面文件"]]
        if not old_file:
            continue
        folder = normalize_rel(str(old_file)).split("\\", 1)[0]
        if category in EXCLUDED_CATEGORIES or folder in EXCLUDED_FOLDERS:
            continue
        result.append({
            "category": category,
            "old_file": normalize_rel(str(old_file)),
            "old_url": row[idx["旧URL"]],
            "old_summary": row[idx["旧页面功能明细"]],
            "status": row[idx["迁移状态"]],
            "new_files": row[idx["新页面文件"]],
            "new_urls": row[idx["新URL"]],
            "new_summary": row[idx["新页面功能明细"]],
            "compare": row[idx["功能比较"]],
            "perm": row[idx["权限说明"]],
        })
    return result


def build_rows() -> List[List[str]]:
    old_api_map = build_old_api_map()
    new_api_map = build_new_api_map()
    abt_routes = build_abt_routes()
    new_page_candidates = build_new_page_candidates(new_api_map, abt_routes)
    pages = load_page_rows()

    data_rows: List[List[str]] = []

    for page in pages:
        page_path = OLD_VIEWS_ROOT / page["old_file"]
        script = extract_script(read_text(page_path))
        named, namespaces = parse_named_imports_v2(script, OLD_VIEWS_ROOT.parent / "api")
        feature_entries: List[dict] = []

        for entry in scan_called_imports(script, named, namespaces):
            api_key = (entry["module"], entry["function"])
            old_meta = old_api_map.get(api_key)
            if not old_meta and entry["module"].endswith(".ts"):
                old_meta = old_api_map.get((entry["module"].replace(".ts", ".js"), entry["function"]))
            if not old_meta and entry["module"].endswith(".js"):
                old_meta = old_api_map.get((entry["module"].replace(".js", ".ts"), entry["function"]))
            ipms_url = old_meta["url_pattern"] if old_meta else ""
            feature_entries.append({
                "action": classify_action(url=ipms_url, fn=entry["function"]),
                "feature": describe_feature(classify_action(url=ipms_url, fn=entry["function"]), page["old_summary"] or "", page["old_file"], url=ipms_url),
                "ipms_call": f"{entry['module']}::{entry['function']}",
                "ipms_url": ipms_url,
                "backend_hint": old_meta["backend_hint"] if old_meta else "",
            })

        for helper in scan_helper_urls(script):
            action = classify_action(url=helper["url"], fn=helper["helper"])
            feature_entries.append({
                "action": action,
                "feature": describe_feature(action, page["old_summary"] or "", page["old_file"], url=helper["url"]),
                "ipms_call": helper["helper"],
                "ipms_url": helper["url"],
                "backend_hint": "Java" if "Java" in helper["helper"] or helper["url"].startswith("/wf/") or helper["url"].startswith("/purchase/") else "C#",
            })

        for route in scan_router_targets(script):
            action = classify_action(route_target=route["target"])
            feature_entries.append({
                "action": action,
                "feature": describe_feature(action, page["old_summary"] or "", page["old_file"], route_target=route["target"]),
                "ipms_call": "router.push",
                "ipms_url": "",
                "backend_hint": "",
                "old_front_target": route["target"],
            })

        dedup = {}
        for item in feature_entries:
            key = (item["action"], item["ipms_call"], item.get("ipms_url", ""), item.get("old_front_target", ""))
            dedup[key] = item
        feature_entries = sorted(dedup.values(), key=lambda x: (ACTION_ORDER.get(x["action"], 99), x["feature"]))

        candidate_rows: List[dict] = []
        for new_file in route_path_to_list(page["new_files"]):
            candidate_rows.extend(new_page_candidates.get(normalize_rel(new_file), []))

        for item in feature_entries:
            matched = choose_new_match(item["action"], candidate_rows, item.get("ipms_url", "")) if page["status"] != "未迁移" else None
            data_rows.append([
                page["category"],
                page["old_file"],
                "；".join(route_path_to_list(page["old_url"])),
                item["action"],
                item["feature"],
                item["ipms_call"],
                item.get("old_front_target", ""),
                item["ipms_url"],
                item["backend_hint"],
                page["status"],
                str(page["new_files"] or ""),
                str(page["new_urls"] or ""),
                matched["new_call"] if matched else "",
                matched["new_api_url"] if matched else "",
                matched["abt_url"] if matched else "",
                matched["controller"] if matched else "",
                PAGE_PRIORITY.get(page["category"], "P3"),
                str(page["compare"] or ""),
            ])
    return data_rows


def build_summary_sheet(wb: Workbook, rows: List[List[str]]) -> None:
    ws = wb.create_sheet("页面汇总")
    header = ["功能分类", "页面数", "功能点数", "已迁移功能点", "部分迁移功能点", "未迁移功能点"]
    ws.append(header)
    counter = defaultdict(lambda: {"pages": set(), "total": 0, "已迁移": 0, "部分迁移": 0, "未迁移": 0})
    for row in rows:
        category, old_file, status = row[0], row[1], row[9]
        counter[category]["pages"].add(old_file)
        counter[category]["total"] += 1
        if status in counter[category]:
            counter[category][status] += 1
    for category in sorted(counter):
        item = counter[category]
        ws.append([category, len(item["pages"]), item["total"], item["已迁移"], item["部分迁移"], item["未迁移"]])
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_note_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("说明")
    notes = [
        ["生成日期", "2026-06-08"],
        ["清单范围", "以 client_prod/src/views 页面为主线，排除低代码、系统管理/系统配置类页面"],
        ["主表粒度", "按 页面 + 功能点 拆行，例如分页查询、导出、删除、详情跳转"],
        ["IPMS后端URL", "来自已整理的 client_prod-url-inventory.csv，按旧前端真实调用接口展示"],
        ["abt-lims后端URL", "优先按 lab-lims-frontend 对应新页面调用 API 反推，再映射到 abt-lims 实际 controller URL；未找到则留空"],
        ["注意", "部分 workflow 页面新前端仍保留旧路径风格，Excel 中已尽量映射到 abt-lims 实际 /api/admin/... controller"],
    ]
    for row in notes:
        ws.append(row)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    widths = {
        "A": 18, "B": 32, "C": 28, "D": 12, "E": 36, "F": 28, "G": 24, "H": 26, "I": 10,
        "J": 10, "K": 32, "L": 28, "M": 28, "N": 24, "O": 34, "P": 46, "Q": 8, "R": 48
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    rows = build_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "页面功能迁移计划"
    header = [
        "功能分类",
        "旧页面文件",
        "旧页面URL",
        "功能动作",
        "功能点说明",
        "IPMS前端调用",
        "旧前端跳转URL",
        "IPMS后端URL",
        "IPMS后端归属",
        "迁移状态",
        "新页面文件",
        "新页面URL",
        "新前端调用",
        "新前端API URL",
        "abt-lims后端URL",
        "abt-lims Controller",
        "优先级",
        "功能差异/备注",
    ]
    ws.append(header)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    build_summary_sheet(wb, rows)
    build_note_sheet(wb)
    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
