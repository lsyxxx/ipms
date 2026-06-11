from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"E:\github\ipms\docs\client-prod-views-page-migration.xlsx")
OUTPUT_PATH = Path(r"E:\github\ipms\docs\client-prod-views-page-migration-detailed.xlsx")
OLD_VIEWS_ROOT = Path(r"E:\github\client_prod\src\views")
NEW_VIEWS_ROOT = Path(r"E:\github\lims\lab-lims-frontend\src\views")


def read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def unique_keep_order(values: list[str]) -> list[str]:
    seen: OrderedDict[str, None] = OrderedDict()
    for value in values:
        value = (value or "").strip()
        if value and value not in seen:
            seen[value] = None
    return list(seen.keys())


def normalize_label(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    value = value.replace("&nbsp;", " ").replace("\n", " ").replace("\r", " ")
    value = re.sub(r"\s+", " ", value).strip(" :：'\"")
    return value


def parse_template(text: str) -> str:
    match = re.search(r"<template>([\s\S]*?)</template>", text, re.IGNORECASE)
    return match.group(1) if match else text


def parse_script(text: str) -> str:
    match = re.search(r"<script[^>]*>([\s\S]*?)</script>", text, re.IGNORECASE)
    return match.group(1) if match else text


def extract_attr_values(text: str, attr_name: str) -> list[str]:
    patterns = [
        rf'{attr_name}="([^"]+)"',
        rf"{attr_name}='([^']+)'",
        rf':{attr_name}="\'([^\"]+)\'"',
        rf":{attr_name}=\"'([^']+)'\"",
    ]
    result: list[str] = []
    for pattern in patterns:
        result.extend(re.findall(pattern, text))
    return unique_keep_order([normalize_label(item) for item in result if normalize_label(item)])


def extract_form_labels(template: str) -> list[str]:
    labels: list[str] = []
    for match in re.finditer(r"<el-form-item\b[^>]*label=\"([^\"]+)\"", template, re.IGNORECASE):
        labels.append(normalize_label(match.group(1)))
    for match in re.finditer(r"<el-form-item\b[^>]*label='([^']+)'", template, re.IGNORECASE):
        labels.append(normalize_label(match.group(1)))
    return unique_keep_order(labels)


def extract_button_texts(template: str) -> list[str]:
    texts: list[str] = []
    for match in re.finditer(r"<el-button\b[^>]*>([\s\S]*?)</el-button>", template, re.IGNORECASE):
        text = normalize_label(match.group(1))
        if text:
            texts.append(text)
    for match in re.finditer(r"<permission-btn\b[^>]*></permission-btn>", template, re.IGNORECASE):
        texts.append("动态权限按钮")
    return unique_keep_order(texts)


def extract_table_labels(template: str) -> list[str]:
    labels: list[str] = []
    for match in re.finditer(r"<el-table-column\b[^>]*label=\"([^\"]+)\"", template, re.IGNORECASE):
        labels.append(normalize_label(match.group(1)))
    for match in re.finditer(r"<el-table-column\b[^>]*label='([^']+)'", template, re.IGNORECASE):
        labels.append(normalize_label(match.group(1)))
    for match in re.finditer(r"<el-table-column\b[^>]*:label=\"'([^']+)'\"", template, re.IGNORECASE):
        labels.append(normalize_label(match.group(1)))
    return [label for label in labels if label not in {"操作"}]


def extract_section_titles(template: str) -> list[str]:
    titles: list[str] = []
    for match in re.finditer(r"<el-tab-pane\b[^>]*label=\"([^\"]+)\"", template, re.IGNORECASE):
        titles.append(normalize_label(match.group(1)))
    for match in re.finditer(r"<span[^>]*>([\s\S]*?)</span>", template, re.IGNORECASE):
        text = normalize_label(match.group(1))
        if 2 <= len(text) <= 12 and any(key in text for key in ["信息", "设置", "选项", "清单", "详情", "样品", "委托", "配置"]):
            titles.append(text)
    return unique_keep_order(titles)


def extract_urls(script: str) -> list[str]:
    urls = re.findall(r"['\"](/[^'\"]+)['\"]", script)
    urls = [u for u in urls if not u.startswith("//")]
    return unique_keep_order(urls)


def extract_router_targets(script: str) -> list[str]:
    targets: list[str] = []
    targets.extend(re.findall(r"\$router\.push\(\s*['\"]([^'\"]+)['\"]", script))
    targets.extend(re.findall(r"path:\s*['\"]([^'\"]+)['\"]", script))
    return unique_keep_order([item for item in targets if item.startswith("/")])


def infer_subject(route_title: str, rel_path: str) -> str:
    stem = Path(rel_path).stem
    title_hints = {
        "purchasingAllList": "采购申请台账",
        "purchasingApplyList": "我的采购申请",
        "entrustnew": "委托单/项目",
        "entrustDetail": "委托项目详情",
        "testTaskView": "检测任务汇总",
        "inventoryList": "库存",
    }
    if stem in title_hints:
        return title_hints[stem]
    if route_title and "模块业务页面" not in route_title:
        title = route_title
        for suffix in ["列表", "台账", "记录", "详情", "查询", "新增", "编辑", "配置", "管理", "汇总", "表单"]:
            if title.endswith(suffix):
                title = title[: -len(suffix)]
        return title or route_title
    folder = Path(rel_path).parent.name
    hints = {
        "weituodan": "委托/项目",
        "workflow": "审批业务",
        "check": "检测项目",
        "standard": "检测标准",
        "instrument": "设备",
        "stock": "库存",
        "settlement": "结算",
        "salary": "工资",
        "fieldWork": "野外作业",
        "safety": "安全检查",
    }
    return hints.get(folder, stem)


def extract_capabilities(template: str, script: str, buttons: list[str]) -> list[str]:
    caps: list[str] = []
    script_lower = script.lower()
    button_blob = " ".join(buttons)
    if "<el-table" in template:
        caps.append("列表展示")
    if "<pagination" in template:
        caps.append("分页查询")
    if "<el-form" in template or "handleFilter" in script:
        caps.append("条件筛选")
    if any(word in button_blob for word in ["新增", "添加", "提交申请"]) or "btnAdd" in script:
        caps.append("新增")
    if any(word in button_blob for word in ["编辑", "修改"]) or "btnEdit" in script:
        caps.append("编辑")
    if any(word in button_blob for word in ["删除", "移除"]) or "btnDel" in script or "delete" in script_lower:
        caps.append("删除")
    if any(word in button_blob for word in ["详情", "查看"]) or any(flag in script for flag in ["/detail", "/dtl/", "btnDetail"]):
        caps.append("详情查看")
    if "导出" in button_blob or any(token in script_lower for token in ["exportexcel", "downloadjavaapi", "purchaseDownload".lower(), ".download ="]):
        caps.append("导出下载")
    if "导入" in button_blob or any(
        token in script_lower
        for token in ["importdialogvisible", "btnbatchimport", "uploadfileexcel", "uploadfileimport", "importentrust", "beforeuploadimport"]
    ):
        caps.append("导入")
    if "<el-upload" in template or "uploadfile" in script_lower:
        caps.append("附件上传")
    if "打印" in button_blob or "print" in script_lower or ".pdf" in script_lower:
        caps.append("打印/PDF")
    if "复制" in button_blob or "btnCopy" in script:
        caps.append("复制")
    if any(word in button_blob for word in ["撤销", "重新提交", "验收", "待我处理", "我已处理"]) or any(
        word in script for word in ["revoke", "restart", "goAccept", "findMyTodos", "findMyDones"]
    ):
        caps.append("审批流操作")
    return unique_keep_order(caps)


def summarize_page(route_title: str, rel_path: str, template: str, script: str) -> tuple[str, str, str]:
    labels = extract_form_labels(template)
    placeholders = extract_attr_values(template, "placeholder")
    buttons = extract_button_texts(template)
    table_labels = extract_table_labels(template)
    section_titles = extract_section_titles(template)
    urls = extract_urls(script)
    router_targets = extract_router_targets(script)
    capabilities = extract_capabilities(template, script, buttons)
    subject = infer_subject(route_title, rel_path)

    parts: list[str] = []
    rel_path_norm = rel_path.replace("/", "\\")
    if rel_path_norm == r"workflow\purchasingAllList.vue":
        return (
            "分页查询采购申请台账；支持按关键字、状态、部门、创建时间筛选；支持导出采购记录、删除记录、点击审批编号查看采购详情；列表展示审批编号、是否验收、总金额、状态、申请人、部门、当前节点、当前审批人、创建时间等字段",
            "列表展示、分页查询、条件筛选、删除、详情查看、导出下载",
            "/purchase/export/list；/purchase/dtl/",
        )
    if rel_path_norm == r"workflow\purchasingApplyList.vue":
        return (
            "分页查询我的采购申请；支持按关键字、状态、创建时间筛选，并按我创建的/待我处理/我已处理切换视图；支持提交申请、导出、撤销、重新提交、验收、点击审批编号查看详情；列表展示审批编号、是否验收、状态、申请人、部门、当前节点、当前审批人、创建时间等字段",
            "列表展示、分页查询、条件筛选、新增、详情查看、导出下载、审批流操作",
            "/purchase/export/list；/purchase/apply；/purchase/accept/；/purchase/dtl/",
        )
    if rel_path_norm == r"weituodan\entrustnew.vue":
        return (
            "分页查询委托单/项目数据；支持按名称/分类编码检索；支持新增、编辑、删除、复制、查看项目详情、导出列表、批量导入委托及样品、下载导入模板；弹窗中可维护委托方信息、样品信息、报告信息并上传附件/样品图片，支持打印项目单和样品 Excel 导入",
            "列表展示、分页查询、条件筛选、新增、编辑、删除、详情查看、导出下载、导入、附件上传、打印/PDF、复制",
            "/weituodan/entrustDetail/；/importEntrustAndSamples；/getImportErrorPage",
        )

    if "分页查询" in capabilities:
        parts.append(f"分页查询{subject}数据")
    elif "列表展示" in capabilities:
        parts.append(f"查询并展示{subject}数据")
    elif any(token in route_title for token in ["详情", "编辑", "新增", "配置", "设置"]):
        parts.append(f"处理{subject}相关信息")

    filter_labels = [item for item in labels if item not in {"", "操作"}][:6]
    if filter_labels:
        parts.append("支持按" + " / ".join(filter_labels) + "筛选或录入")
    elif placeholders:
        parts.append("支持按" + " / ".join(placeholders[:4]) + "检索")

    visible_actions: list[str] = []
    action_map = {
        "新增": "新增",
        "编辑": "编辑",
        "删除": "删除",
        "详情查看": "查看详情",
        "导出下载": "导出/下载",
        "导入": "导入",
        "附件上传": "附件上传",
        "打印/PDF": "打印/PDF导出",
        "复制": "复制",
        "审批流操作": "审批流处理",
        "条件筛选": "条件筛选",
    }
    for capability in capabilities:
        if capability in action_map:
            visible_actions.append(action_map[capability])
    if visible_actions:
        parts.append("支持" + "、".join(visible_actions))

    if table_labels:
        parts.append("列表展示" + "、".join(table_labels[:8]) + "等字段")
    if section_titles:
        parts.append("页面包含" + "、".join(section_titles[:6]) + "等区块")
    if router_targets:
        parts.append("可跳转到" + "、".join(router_targets[:5]))

    summary = "；".join(unique_keep_order(parts)) if parts else (route_title or rel_path)

    top_urls = []
    for url in urls:
        if url not in top_urls:
            top_urls.append(url)
        if len(top_urls) >= 8:
            break
    url_summary = "；".join(top_urls)

    return summary, "、".join(capabilities), url_summary


def compare_features(status: str, old_caps: list[str], new_caps: list[str], new_files: str) -> str:
    old_set = set(old_caps)
    new_set = set(new_caps)
    covered = [item for item in old_caps if item in new_set]
    missing = [item for item in old_caps if item not in new_set]
    if status == "已迁移":
        if missing:
            return "核心功能已迁移，当前新页明显覆盖" + "、".join(covered[:6]) + "；但未明显体现" + "、".join(missing[:4])
        return "核心功能已迁移，旧页能力在新页基本可完成"
    if status == "部分迁移":
        if new_files:
            target_desc = "新系统拆分到" + new_files.replace("；", "、")
        else:
            target_desc = "新系统已有部分承接页面"
        if covered and missing:
            return target_desc + "；已覆盖" + "、".join(covered[:6]) + "；未明显覆盖" + "、".join(missing[:4])
        return target_desc + "；仅覆盖部分流程"
    if status == "未迁移":
        return "旧页包含" + "、".join(old_caps[:6]) + "等能力，新前端暂未落地对应页面"
    return "本页不纳入本次迁移范围"


def split_paths(cell_value: str) -> list[str]:
    if not cell_value:
        return []
    parts = re.split(r"[；;]", str(cell_value))
    return [item.strip() for item in parts if item and item.strip().lower() != "none"]


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["views页面迁移清单"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    header_map = {value: index + 1 for index, value in enumerate(headers) if value}

    extra_headers = [
        "旧页面功能明细",
        "旧页面能力标签",
        "旧页面主要接口",
        "新页面功能明细",
        "新页面能力标签",
        "新页面主要接口",
    ]
    for header in extra_headers:
        if header not in header_map:
            ws.cell(row=1, column=ws.max_column + 1, value=header)
            header_map[header] = ws.max_column

    for row_idx in range(2, ws.max_row + 1):
        old_rel = ws.cell(row=row_idx, column=header_map["旧页面文件"]).value
        old_title = ws.cell(row=row_idx, column=header_map["旧页面标题/功能"]).value or ""
        status = ws.cell(row=row_idx, column=header_map["迁移状态"]).value or ""
        new_files_cell = ws.cell(row=row_idx, column=header_map["新页面文件"]).value or ""

        old_text = read_text(OLD_VIEWS_ROOT / str(old_rel)) if old_rel else ""
        old_template = parse_template(old_text)
        old_script = parse_script(old_text)
        old_summary, old_caps_text, old_url_text = summarize_page(str(old_title), str(old_rel), old_template, old_script)
        old_caps = [item for item in old_caps_text.split("、") if item]

        new_summaries: list[str] = []
        new_caps: list[str] = []
        new_urls: list[str] = []
        for new_rel in split_paths(str(new_files_cell)):
            new_text = read_text(NEW_VIEWS_ROOT / new_rel)
            if not new_text:
                continue
            new_template = parse_template(new_text)
            new_script = parse_script(new_text)
            summary, caps_text, url_text = summarize_page("", new_rel, new_template, new_script)
            if summary:
                new_summaries.append(f"{new_rel}: {summary}")
            new_caps.extend([item for item in caps_text.split("、") if item])
            if url_text:
                new_urls.append(f"{new_rel}: {url_text}")

        new_caps = unique_keep_order(new_caps)
        compare_text = compare_features(str(status), old_caps, new_caps, str(new_files_cell))

        ws.cell(row=row_idx, column=header_map["旧页面功能明细"], value=old_summary)
        ws.cell(row=row_idx, column=header_map["旧页面能力标签"], value=old_caps_text)
        ws.cell(row=row_idx, column=header_map["旧页面主要接口"], value=old_url_text)
        ws.cell(
            row=row_idx,
            column=header_map["新页面功能明细"],
            value="；".join(new_summaries) if new_summaries else "",
        )
        ws.cell(row=row_idx, column=header_map["新页面能力标签"], value="、".join(new_caps))
        ws.cell(row=row_idx, column=header_map["新页面主要接口"], value="；".join(new_urls))
        ws.cell(row=row_idx, column=header_map["功能比较"], value=compare_text)
        if old_summary:
            ws.cell(row=row_idx, column=header_map["旧页面标题/功能"], value=old_summary)
        if new_summaries:
            ws.cell(row=row_idx, column=header_map["新页面/功能说明"], value="；".join(new_summaries))

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()
